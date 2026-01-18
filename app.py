import streamlit as st
import pandas as pd
import json
import io
import os
import warnings
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from datetime import datetime

# Page setup (Must be the first line)
st.set_page_config(page_title="Personal Budget Analyzer", page_icon="💰", layout="wide")

# Silence warnings
warnings.filterwarnings('ignore')

# ==========================================
# 1. Logic & Helpers
# ==========================================

DEFAULT_CATEGORY_MAPPING = {
    "Groceries & Supermarket": ["רמי לוי", "שופרסל", "AM:PM"],
    "Uncategorized": []
}


@st.cache_data  # Cache data to avoid reloading on every interaction
def load_category_mapping():
    try:
        if os.path.exists('categories.json'):
            with open('categories.json', 'r', encoding='utf-8') as f:
                return json.load(f)
    except Exception:
        pass
    return DEFAULT_CATEGORY_MAPPING


def classify_expense(merchant_name, mapping):
    if not isinstance(merchant_name, str):
        return "Uncategorized"
    merchant_upper = merchant_name.upper()
    for category, keywords in mapping.items():
        for keyword in keywords:
            if keyword.upper() in merchant_upper:
                return category
    return "Uncategorized"


def clean_dates(date_series):
    date_series = date_series.astype(str).str.strip()
    date_series = date_series.str.replace('.', '-', regex=False)
    date_series = date_series.str.replace('/', '-', regex=False)
    return pd.to_datetime(date_series, dayfirst=True, errors='coerce')


def read_excel_file(uploaded_file):
    """
    Streamlit-adapted function that accepts a file object instead of a path.
    """
    try:
        # Initial read to detect the header row
        preview = pd.read_excel(uploaded_file, header=None, nrows=30)

        header_row = 0
        for i, row in preview.iterrows():
            row_str = row.astype(str).str.cat(sep=' ')
            # Keep Hebrew keywords here to match the input files
            if "תאריך" in row_str and ("רכישה" in row_str or "עסקה" in row_str):
                header_row = i
                break

        # Reset file pointer to the beginning to read again
        uploaded_file.seek(0)
        df = pd.read_excel(uploaded_file, skiprows=header_row)
        df.columns = df.columns.str.strip()

        # Dynamic column detection (Hebrew headers)
        merchant_col = next((c for c in ['שם בית עסק', 'שם בית העסק'] if c in df.columns), None)
        amount_col = 'סכום חיוב' if 'סכום חיוב' in df.columns else None
        date_col = next((c for c in ['תאריך רכישה', 'תאריך עסקה', 'תאריך'] if c in df.columns), None)

        if not merchant_col or not amount_col:
            return None

        temp_df = pd.DataFrame()
        temp_df['Date'] = clean_dates(df[date_col]) if date_col else pd.NaT
        temp_df['Merchant'] = df[merchant_col]
        temp_df['Amount'] = df[amount_col]
        temp_df['Source_File'] = uploaded_file.name

        # Clean summary rows
        summary_phrases = ['TOTAL FOR DATE', 'סה"כ לחיוב', 'סה"כ', 'סך הכל']
        mask = ~temp_df['Merchant'].astype(str).str.contains('|'.join(summary_phrases), regex=True, na=False)
        temp_df = temp_df[mask]

        temp_df = temp_df.dropna(subset=['Amount'])
        temp_df['Amount'] = temp_df['Amount'].astype(str).str.replace(r'[₪,]', '', regex=True).str.strip()
        temp_df['Amount'] = pd.to_numeric(temp_df['Amount'], errors='coerce')
        temp_df = temp_df.dropna(subset=['Amount'])

        return temp_df

    except Exception as e:
        st.error(f"Error reading file {uploaded_file.name}: {e}")
        return None


# ==========================================
# 2. Excel Generation Logic (Memory Buffer)
# ==========================================

def format_excel_sheet(worksheet):
    # Styles
    header_fill = PatternFill(start_color="D9E1F2", fill_type="solid")
    total_col_fill = PatternFill(start_color="FFF2CC", fill_type="solid")
    grand_total_fill = PatternFill(start_color="E2EFDA", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    bold_font = Font(bold=True)

    # Freeze panes (Row 1 and Column A)
    worksheet.freeze_panes = 'B2'

    max_row = worksheet.max_row
    max_col = worksheet.max_column

    for row in worksheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.border = thin_border
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00 "₪"'

            if cell.row == 1:
                cell.fill = header_fill
                cell.font = bold_font
                cell.alignment = Alignment(horizontal='center')
            elif cell.row == max_row:
                cell.fill = grand_total_fill
                cell.font = bold_font
            elif cell.column == max_col:
                cell.fill = total_col_fill
                cell.font = bold_font

    # Auto-fit column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    cell_len = len(str(cell.value))
                    if cell_len > max_length:
                        max_length = cell_len
            except:
                pass
        worksheet.column_dimensions[column_letter].width = (max_length + 2) * 1.2


def generate_excel_bytes(full_df, pivot_table):
    """Generates the Excel file in memory (BytesIO) for download."""
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        full_df.to_excel(writer, sheet_name='Transactions', index=False)
        format_excel_sheet(writer.sheets['Transactions'])

        pivot_table.to_excel(writer, sheet_name='Summary')
        format_excel_sheet(writer.sheets['Summary'])

    output.seek(0)
    return output


# ==========================================
# 3. User Interface (Streamlit UI)
# ==========================================

st.title("📊 Smart Expense Analysis")
st.markdown("Drag and drop your bank or credit card Excel files, and the system will do the rest.")

# Sidebar
with st.sidebar:
    st.header("Settings")
    mapping = load_category_mapping()
    st.success(f"Loaded {len(mapping)} categories from JSON")
    st.markdown("---")
    st.info("Tip: Ensure 'categories.json' is in the same folder.")

# Step 1: File Upload
uploaded_files = st.file_uploader("Upload Excel files (Multi-select supported)", type=['xlsx', 'xls'],
                                  accept_multiple_files=True)

if uploaded_files:
    all_data = []
    for file in uploaded_files:
        df = read_excel_file(file)
        if df is not None:
            all_data.append(df)

    if all_data:
        # Merge all files
        full_df = pd.concat(all_data, ignore_index=True)

        # Sort and Classify
        full_df = full_df.sort_values('Date', ascending=False)
        full_df['Month_Year'] = full_df['Date'].dt.to_period('M')
        full_df['Category'] = full_df['Merchant'].apply(lambda x: classify_expense(x, mapping))

        # --- Visual Dashboard ---

        # Key Metrics
        total_spent = full_df['Amount'].sum()
        col1, col2, col3 = st.columns(3)
        col1.metric("Total Expenses", f"₪{total_spent:,.2f}")
        col2.metric("Total Transactions", len(full_df))
        col3.metric("Active Months", full_df['Month_Year'].nunique())

        st.divider()

        # Chart by Category
        st.subheader("Distribution by Category")
        category_sum = full_df.groupby('Category')['Amount'].sum().sort_values(ascending=False)
        st.bar_chart(category_sum)

        # Pivot Table (Logic for Excel)
        pivot_table = full_df.pivot_table(
            index='Category',
            columns='Month_Year',
            values='Amount',
            aggfunc='sum',
            fill_value=0
        )
        pivot_table['TOTAL'] = pivot_table.sum(axis=1)
        pivot_table.loc['GRAND TOTAL'] = pivot_table.sum()

        # Raw Data Preview
        with st.expander("Show Raw Data"):
            st.dataframe(full_df)

        st.divider()

        # Download Button
        excel_data = generate_excel_bytes(full_df, pivot_table)

        st.download_button(
            label="📥 Download Formatted Excel Report",
            data=excel_data,
            file_name=f'Expenses_Report_{datetime.strftime(datetime.today(), "%d_%m_%Y")}.xlsx',
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary"  # Prominent button
        )

    else:
        st.warning("No transactions found in the uploaded files.")
import io

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def format_excel_sheet(worksheet):
    header_fill = PatternFill(start_color="D9E1F2", fill_type="solid")
    total_col_fill = PatternFill(start_color="FFF2CC", fill_type="solid")
    grand_total_fill = PatternFill(start_color="E2EFDA", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    bold_font = Font(bold=True)

    worksheet.freeze_panes = 'B2'

    max_row = worksheet.max_row
    max_col = worksheet.max_column

    for row in worksheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.border = thin_border
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00 "₪"'

            if cell.row == 1:
                cell.fill = header_fill
                cell.font = bold_font
                cell.alignment = Alignment(horizontal='center')
            elif cell.row == max_row:
                cell.fill = grand_total_fill
                cell.font = bold_font
            elif cell.column == max_col:
                cell.fill = total_col_fill
                cell.font = bold_font

    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    cell_len = len(str(cell.value))
                    if cell_len > max_length:
                        max_length = cell_len
            except Exception:
                pass
        worksheet.column_dimensions[column_letter].width = (max_length + 2) * 1.2


def generate_excel_bytes(full_df, pivot_table):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        full_df.to_excel(writer, sheet_name='Transactions', index=False)
        format_excel_sheet(writer.sheets['Transactions'])

        pivot_table.to_excel(writer, sheet_name='Summary')
        format_excel_sheet(writer.sheets['Summary'])

    output.seek(0)
    return output
